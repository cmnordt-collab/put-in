#!/usr/bin/env python3
"""
Turn the Put-In spreadsheet into the JSON file the website reads.

Run this every time you edit put-in-v1-data.xlsx:

    python3 make_spots.py

It writes site/data/spots.json. Nothing else needs touching.
"""

import json, os, datetime
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "put-in-v1-data.xlsx")
OUT = os.path.join(HERE, "site", "data", "spots.json")

BLANK = {"", "unknown", "n/a", "na", "none recorded", "tbd"}


def clean(v):
    """Empty and 'unknown' both become None, so the site can say 'not recorded'
    rather than pretending absence is a No."""
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    return None if s.lower() in BLANK else s


def main():
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["Spot data"]
    headers = [c.value for c in ws[1]]

    spots = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = {}
        for key, val in zip(headers, row):
            c = clean(val)
            if c is not None:
                rec[key] = c
        if rec.get("spot_id") and rec.get("name"):
            spots.append(rec)

    spots.sort(key=lambda r: r["name"])

    payload = {
        "generated": datetime.date.today().isoformat(),
        "count": len(spots),
        "note": ("A blank field means nobody has verified it yet. It does not mean no. "
                 "This matters most for hazards and dams."),
        "sources": ["NC State Parks", "Lower Haw River State Natural Area",
                    "NC Wildlife Resources Commission", "Durham Parks and Recreation",
                    "Haw River Assembly", "Eno River Association"],
        "spots": spots,
    }

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=1, ensure_ascii=False)

    filled = {}
    for s in spots:
        for k in s:
            filled[k] = filled.get(k, 0) + 1

    print(f"wrote {len(spots)} spots to {OUT}")
    print(f"  with coordinates : {filled.get('latitude', 0)}")
    print(f"  with a USGS gage : {filled.get('usgs_gage_id', 0)}")
    print(f"  with hazards noted: {filled.get('hazards', 0)}")
    print(f"  with walk to water: {filled.get('walk_to_water', 0)}")


if __name__ == "__main__":
    main()
